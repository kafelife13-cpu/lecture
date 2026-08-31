import json
import re
import sys
import urllib.parse
import urllib.request

import openpyxl

SB_URL = "https://zniwzwanlvvmzkkcmikh.supabase.co"
SB_KEY = "sb_publishable_IbfVcXs5TmXfumrnK4Somg_7MzygCdh"


def phone(value):
    digits = re.sub(r"\D", "", str(value or ""))
    return digits if re.fullmatch(r"01\d{8,9}", digits) else ""


def main(path):
    book = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet = book.active
    by_id = {}
    duplicates = set()
    for row in range(4, sheet.max_row + 1):
        student_phone = phone(sheet.cell(row, 14).value)  # N: 학생 연락처
        parent_phone = phone(sheet.cell(row, 16).value)   # P: 학부모 연락처
        if not student_phone:
            continue
        student_id = student_phone[3:7]
        if student_id in by_id:
            duplicates.add(student_id)
            continue
        by_id[student_id] = {"student_phone": student_phone, "parent_phone": parent_phone or None}

    headers = {"apikey": SB_KEY, "Authorization": f"Bearer {SB_KEY}", "Content-Type": "application/json"}
    query = urllib.parse.urlencode({"select": "id,name,role,status", "role": "eq.student"})
    request = urllib.request.Request(f"{SB_URL}/rest/v1/users?{query}", headers=headers)
    with urllib.request.urlopen(request, timeout=30) as response:
        users = json.load(response)
    matched = []
    unmatched = []
    for user in users:
        values = by_id.get(str(user["id"]))
        if not values:
            unmatched.append(user["name"])
            continue
        patch_query = urllib.parse.urlencode({"id": f"eq.{user['id']}"})
        request = urllib.request.Request(
            f"{SB_URL}/rest/v1/users?{patch_query}",
            headers={**headers, "Prefer": "return=minimal"},
            data=json.dumps(values).encode("utf-8"),
            method="PATCH",
        )
        with urllib.request.urlopen(request, timeout=30):
            pass
        matched.append(user["name"])

    print(json.dumps({
        "matched_count": len(matched),
        "unmatched_count": len(unmatched),
        "unmatched_names": unmatched,
        "duplicate_ids": sorted(duplicates),
    }, ensure_ascii=False))


if __name__ == "__main__":
    main(sys.argv[1])

